from decimal import Decimal
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Product, ProductGroup

PRODUCT_UID_KEYS = ['Унікальний_ідентифікатор', 'unique_id', 'uid']
PRODUCT_NAME_KEYS = ['Назва_позиції', 'name', 'Название_позиции']
PRODUCT_PRICE_KEYS = ['Ціна', 'price', 'Цена']
PRODUCT_QTY_KEYS = ['Кількість', 'quantity', 'Количество']
PRODUCT_AVAIL_KEYS = ['Наявність', 'availability', 'Наличие']
PRODUCT_GROUP_KEYS = ['ID_групи', 'group_id', 'Идентификатор_группы']

GROUP_UID_KEYS = ['Унікальний_ідентифікатор', 'group_uid', 'id']
GROUP_NAME_KEYS = ['Назва', 'name', 'Название']


def _normalize_headers(row):
    return {str(cell).strip(): idx for idx, cell in enumerate(row) if cell is not None and str(cell).strip()}


def _pick(headers: dict[str, int], keys: list[str]) -> int | None:
    for k in keys:
        if k in headers:
            return headers[k]
    return None


def _to_int(value, default=0):
    if value is None or value == '':
        return default
    try:
        return int(float(value))
    except Exception:
        return default


def _to_decimal(value, default='0'):
    if value is None or value == '':
        return Decimal(default)
    try:
        return Decimal(str(value).replace(',', '.'))
    except Exception:
        return Decimal(default)


def import_prom_xlsx(db: Session, content: bytes) -> dict[str, int]:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Invalid XLSX file') from exc

    if len(wb.sheetnames) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='XLSX must have at least 2 sheets: products and groups')

    products_ws = wb[wb.sheetnames[0]]
    groups_ws = wb[wb.sheetnames[1]]

    groups_rows = list(groups_ws.iter_rows(values_only=True))
    products_rows = list(products_ws.iter_rows(values_only=True))
    if not groups_rows or not products_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Sheets are empty')

    group_headers = _normalize_headers(groups_rows[0])
    g_uid_idx = _pick(group_headers, GROUP_UID_KEYS)
    g_name_idx = _pick(group_headers, GROUP_NAME_KEYS)
    if g_uid_idx is None or g_name_idx is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Groups sheet columns are invalid')

    upserted_groups = 0
    group_map: dict[str, int] = {}
    for row in groups_rows[1:]:
        uid_raw = row[g_uid_idx] if g_uid_idx < len(row) else None
        name_raw = row[g_name_idx] if g_name_idx < len(row) else None
        if uid_raw is None or name_raw is None:
            continue
        uid = str(uid_raw).strip()
        if not uid:
            continue
        name = str(name_raw).strip()
        existing = db.scalar(select(ProductGroup).where(ProductGroup.prom_uid == uid))
        if existing:
            existing.name = name
            group = existing
        else:
            group = ProductGroup(prom_uid=uid, name=name)
            db.add(group)
        db.flush()
        group_map[uid] = group.id
        upserted_groups += 1

    product_headers = _normalize_headers(products_rows[0])
    p_uid_idx = _pick(product_headers, PRODUCT_UID_KEYS)
    p_name_idx = _pick(product_headers, PRODUCT_NAME_KEYS)
    p_price_idx = _pick(product_headers, PRODUCT_PRICE_KEYS)
    p_qty_idx = _pick(product_headers, PRODUCT_QTY_KEYS)
    p_avail_idx = _pick(product_headers, PRODUCT_AVAIL_KEYS)
    p_group_idx = _pick(product_headers, PRODUCT_GROUP_KEYS)

    if p_uid_idx is None or p_name_idx is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Products sheet columns are invalid')

    upserted_products = 0
    for row in products_rows[1:]:
        uid_raw = row[p_uid_idx] if p_uid_idx < len(row) else None
        if uid_raw is None:
            continue
        uid = str(uid_raw).strip()
        if not uid:
            continue

        name_raw = row[p_name_idx] if p_name_idx is not None and p_name_idx < len(row) else ''
        price_raw = row[p_price_idx] if p_price_idx is not None and p_price_idx < len(row) else 0
        qty_raw = row[p_qty_idx] if p_qty_idx is not None and p_qty_idx < len(row) else 0
        avail_raw = row[p_avail_idx] if p_avail_idx is not None and p_avail_idx < len(row) else 'available'
        group_uid_raw = row[p_group_idx] if p_group_idx is not None and p_group_idx < len(row) else None

        group_id = None
        if group_uid_raw is not None:
            group_id = group_map.get(str(group_uid_raw).strip())

        existing = db.scalar(select(Product).where(Product.prom_uid == uid))
        if existing:
            existing.name = str(name_raw).strip() or existing.name
            existing.price = _to_decimal(price_raw)
            existing.qty = _to_int(qty_raw)
            existing.availability = str(avail_raw).strip() or 'available'
            existing.group_id = group_id
        else:
            db.add(
                Product(
                    prom_uid=uid,
                    name=str(name_raw).strip() or uid,
                    price=_to_decimal(price_raw),
                    qty=_to_int(qty_raw),
                    availability=str(avail_raw).strip() or 'available',
                    group_id=group_id,
                )
            )
        upserted_products += 1

    db.commit()
    return {'groups': upserted_groups, 'products': upserted_products}
